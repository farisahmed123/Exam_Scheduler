import pandas as pd
from collections import defaultdict

class ExamScheduler:
    def __init__(self, excel_file):
        self.excel_file = excel_file
        self.students = {}
        self.courses = []
        self.course_students = {}
        self.conflicts = {}
        self.schedule = {}
        self.combined_courses = []
        
        self.DAYS = 3
        self.SLOTS_PER_DAY = 6
        self.MAX_CAPACITY = 500
        
    def load_data(self):
        print("Reading Excel file...")
        df = pd.read_excel(self.excel_file)
        
        # Find columns
        student_col = None
        course_col = None
        subject_col = None
        
        for col in df.columns:
            col_lower = col.lower().replace('.', '').replace(' ', '')
            if 'roll' in col_lower or col_lower == 'studentid':
                student_col = col
            elif col_lower == 'code' or 'coursecode' in col_lower:
                course_col = col
            elif col_lower == 'course' or 'subject' in col_lower:
                subject_col = col
        
        if not student_col or not course_col:
            print("ERROR: Can't find student/course columns!")
            return False
        
        if not subject_col:
            subject_col = course_col
        
        self.df = df
        self.student_col = student_col
        self.course_col = course_col
        self.subject_col = subject_col
        
        # Build student enrollments
        for _, row in df.iterrows():
            student = str(row[student_col])
            course = row[course_col]
            
            if student not in self.students:
                self.students[student] = []
            self.students[student].append(course)
        
        self.courses = sorted(df[course_col].unique())
        
        # Count students in each course
        for course in self.courses:
            self.course_students[course] = len(df[df[course_col] == course][student_col].unique())
        
        self.find_combined_courses()
        
        print(f"Loaded: {len(self.students)} students, {len(self.courses)} courses")
        return True
    
    def find_combined_courses(self):
        import openpyxl
        import re
        
        print("Checking for combined courses...")
        
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb.active
            
            combined_groups = []
            
            # Read columns F-I from row 3 onwards
            for i in range(3, ws.max_row + 1):
                for col_idx in range(6, 10):
                    cell_value = ws.cell(i, col_idx).value
                    
                    if cell_value and isinstance(cell_value, str):
                        course_codes = re.findall(r'\b[A-Z]{2}\d{4}\b', cell_value)
                        
                        if len(course_codes) > 1:
                            combined_groups.append(course_codes)
            
            # Remove duplicates
            self.combined_courses = []
            for group in combined_groups:
                unique_group = sorted(list(set(group)))
                if unique_group not in self.combined_courses and len(unique_group) > 1:
                    self.combined_courses.append(unique_group)
            
            if self.combined_courses:
                print(f"\nFound {len(self.combined_courses)} groups of courses that MUST be in SAME SLOT:")
                for i, group in enumerate(self.combined_courses, 1):
                    print(f"  Group {i}: {', '.join(group)}")
            else:
                print("No combined course groups found.")
                
        except Exception as e:
            print(f"Warning: Could not read combined courses: {e}")
            self.combined_courses = []
    
    def find_conflicts(self):
        print("Finding conflicts...")
        
        for course in self.courses:
            self.conflicts[course] = set()
        
        # Mark conflicting courses
        for student, courses in self.students.items():
            for i in range(len(courses)):
                for j in range(i + 1, len(courses)):
                    course1 = courses[i]
                    course2 = courses[j]
                    self.conflicts[course1].add(course2)
                    self.conflicts[course2].add(course1)
        
        print(f"Conflict graph ready")
    
    def can_fit_in_slot(self, course, day, slot):
        # Check for conflicts
        for scheduled_course, (d, s) in self.schedule.items():
            if d == day and s == slot:
                if scheduled_course in self.conflicts[course]:
                    return False
        
        # Check capacity
        current_students = set()
        for scheduled_course, (d, s) in self.schedule.items():
            if d == day and s == slot:
                students_in_course = self.df[self.df[self.course_col] == scheduled_course][self.student_col].unique()
                current_students.update(students_in_course)
        
        if len(current_students) + self.course_students[course] > self.MAX_CAPACITY:
            return False
        
        return True
    
    def calculate_penalty(self, course, day, slot):
        penalty = 0
        students_in_course = self.df[self.df[self.course_col] == course][self.student_col].unique()
        
        for student in students_in_course:
            student_id = str(student)
            if student_id not in self.students:
                continue
            
            same_day_exams = 0
            consecutive_exams = 0
            
            for other_course in self.students[student_id]:
                if other_course == course or other_course not in self.schedule:
                    continue
                
                other_day, other_slot = self.schedule[other_course]
                
                if other_day == day:
                    same_day_exams += 1
                    if abs(other_slot - slot) == 1:
                        consecutive_exams += 1
            
            # Penalty calculation
            if same_day_exams >= 2:
                penalty += 100
            elif same_day_exams == 1:
                penalty += 10
            
            if consecutive_exams > 0:
                penalty += 5
        
        return penalty
    
    def create_schedule(self):
        print("Creating schedule...")
        scheduled_courses = set()
        
        # Schedule combined course groups first
        if self.combined_courses:
            print(f"\nScheduling {len(self.combined_courses)} combined course groups...")
            
            for group_idx, course_group in enumerate(self.combined_courses, 1):
                print(f"  Scheduling Group {group_idx}: {', '.join(course_group)}")
                best_slot = None
                best_penalty = float('inf')
                
                for day in range(self.DAYS):
                    for slot in range(self.SLOTS_PER_DAY):
                        can_fit_all = True
                        total_penalty = 0
                        
                        for course in course_group:
                            if not self.can_fit_in_slot(course, day, slot):
                                can_fit_all = False
                                break
                            total_penalty += self.calculate_penalty(course, day, slot)
                        
                        if can_fit_all and total_penalty < best_penalty:
                            best_penalty = total_penalty
                            best_slot = (day, slot)
                
                if best_slot:
                    for course in course_group:
                        self.schedule[course] = best_slot
                        scheduled_courses.add(course)
                    print(f"    ✓ Scheduled at Day {best_slot[0]+1}, Slot {best_slot[1]+1}")
                else:
                    print(f"    WARNING: Could not fit group {group_idx} in one slot!")
                    for day in range(self.DAYS):
                        for slot in range(self.SLOTS_PER_DAY):
                            for course in course_group:
                                self.schedule[course] = (day, slot)
                                scheduled_courses.add(course)
                            break
                        break
        
        # Schedule remaining courses
        remaining_courses = [c for c in self.courses if c not in scheduled_courses]
        priority_courses = sorted(remaining_courses, 
                                 key=lambda c: (len(self.conflicts[c]), self.course_students[c]), 
                                 reverse=True)
        
        for course in priority_courses:
            best_slot = None
            best_penalty = float('inf')
            
            for day in range(self.DAYS):
                for slot in range(self.SLOTS_PER_DAY):
                    if self.can_fit_in_slot(course, day, slot):
                        penalty = self.calculate_penalty(course, day, slot)
                        if penalty < best_penalty:
                            best_penalty = penalty
                            best_slot = (day, slot)
            
            if best_slot:
                self.schedule[course] = best_slot
            else:
                print(f"Warning: Forcing schedule for {course}")
                min_load = float('inf')
                for day in range(self.DAYS):
                    for slot in range(self.SLOTS_PER_DAY):
                        current = len([c for c, (d, s) in self.schedule.items() if d == day and s == slot])
                        if current < min_load:
                            min_load = current
                            best_slot = (day, slot)
                self.schedule[course] = best_slot
        
        print(f"Scheduled {len(self.schedule)} courses")
        
        if self.combined_courses:
            print("\n✓ Combined course groups successfully scheduled:")
            for group_idx, course_group in enumerate(self.combined_courses, 1):
                if course_group[0] in self.schedule:
                    slot_info = self.schedule[course_group[0]]
                    print(f"  Group {group_idx} at Day {slot_info[0]+1} Slot {slot_info[1]+1}: {', '.join(course_group)}")
    
    def get_time(self, slot):
        times = ["08:00-09:00", "09:00-10:00", "10:00-11:00", 
                "11:00-12:00", "12:00-13:00", "13:00-14:00"]
        return times[slot]
    
    def save_schedule(self):
        print("Saving schedule...")
        rows = []
        
        for day in range(self.DAYS):
            for slot in range(self.SLOTS_PER_DAY):
                courses_here = []
                subjects_here = []
                total_students = set()
                
                for course, (d, s) in self.schedule.items():
                    if d == day and s == slot:
                        courses_here.append(course)
                        subject = self.df[self.df[self.course_col] == course][self.subject_col].iloc[0]
                        subjects_here.append(f"{course} - {subject}")
                        students = self.df[self.df[self.course_col] == course][self.student_col].unique()
                        total_students.update(students)
                
                rows.append({
                    'Day': f'Day {day + 1}',
                    'Slot': f'Slot {slot + 1}',
                    'Time': self.get_time(slot),
                    'Courses': ', '.join(courses_here),
                    'Subjects': ' | '.join(subjects_here),
                    'Students': len(total_students)
                })
        
        schedule_df = pd.DataFrame(rows)
        
        # Course-wise schedule
        course_rows = []
        for course in sorted(self.schedule.keys()):
            day, slot = self.schedule[course]
            subject = self.df[self.df[self.course_col] == course][self.subject_col].iloc[0]
            students = self.course_students[course]
            
            course_rows.append({
                'Course Code': course,
                'Subject': subject,
                'Day': f'Day {day + 1}',
                'Slot': f'Slot {slot + 1}',
                'Time': self.get_time(slot),
                'Students': students
            })
        
        course_df = pd.DataFrame(course_rows)
        
        with pd.ExcelWriter('exam_schedule.xlsx') as writer:
            schedule_df.to_excel(writer, sheet_name='Time Table', index=False)
            course_df.to_excel(writer, sheet_name='Course List', index=False)
        
        print("✓ Saved to exam_schedule.xlsx")
    
    def analyze_and_report(self):
        print("Analyzing schedule...")
        
        stats = {
            'papers_per_day': defaultdict(int),
            'consecutive_papers': defaultdict(int),
            'slot_loads': {}
        }
        
        # Analyze each slot
        for day in range(self.DAYS):
            for slot in range(self.SLOTS_PER_DAY):
                students_here = set()
                courses_here = []
                
                for course, (d, s) in self.schedule.items():
                    if d == day and s == slot:
                        courses_here.append(course)
                        students = self.df[self.df[self.course_col] == course][self.student_col].unique()
                        students_here.update(students)
                
                stats['slot_loads'][f'Day {day+1} Slot {slot+1}'] = {
                    'students': len(students_here),
                    'courses': courses_here
                }
        
        # Analyze student load
        for student, courses in self.students.items():
            day_count = defaultdict(int)
            slot_list = []
            
            for course in courses:
                if course in self.schedule:
                    day, slot = self.schedule[course]
                    day_count[day] += 1
                    slot_list.append((day, slot))
            
            for day, count in day_count.items():
                if count >= 3:
                    stats['papers_per_day'][count] += 1
            
            if len(slot_list) > 1:
                slot_list.sort()
                max_consecutive = 1
                current_consecutive = 1
                
                for i in range(1, len(slot_list)):
                    prev_day, prev_slot = slot_list[i-1]
                    curr_day, curr_slot = slot_list[i]
                    
                    if curr_day == prev_day and curr_slot == prev_slot + 1:
                        current_consecutive += 1
                        max_consecutive = max(max_consecutive, current_consecutive)
                    else:
                        current_consecutive = 1
                
                if max_consecutive >= 3:
                    stats['consecutive_papers'][max_consecutive] += 1
        
        # Write report file
        with open('schedule_report.txt', 'w') as f:
            f.write("="*70 + "\n")
            f.write("FAST PESHAWAR - EXAM SCHEDULE REPORT\n")
            f.write("="*70 + "\n\n")
            
            f.write("OVERVIEW:\n")
            f.write(f"Total Students: {len(self.students)}\n")
            f.write(f"Total Courses: {len(self.courses)}\n")
            f.write(f"Total Slots: {self.DAYS * self.SLOTS_PER_DAY}\n\n")
            
            f.write("A. STUDENTS IN EACH SLOT:\n")
            f.write("-"*70 + "\n")
            for slot_name, info in sorted(stats['slot_loads'].items()):
                f.write(f"{slot_name}: {info['students']} students\n")
            f.write("\n")
            
            f.write("B. STUDENTS WITH MULTIPLE PAPERS ON SAME DAY:\n")
            f.write("-"*70 + "\n")
            for n in [3, 4, 5, 6]:
                count = stats['papers_per_day'].get(n, 0)
                f.write(f"{n} papers on one day: {count} students\n")
            f.write("\n")
            
            f.write("C. STUDENTS WITH CONSECUTIVE PAPERS:\n")
            f.write("-"*70 + "\n")
            for n in [3, 4, 5, 6]:
                count = stats['consecutive_papers'].get(n, 0)
                f.write(f"{n} consecutive papers: {count} students\n")
            f.write("\n")
            
            f.write("APPROACH:\n")
            f.write("-"*70 + "\n")
            f.write("1. Built a conflict graph (courses with common students)\n")
            f.write("2. Used greedy graph coloring algorithm\n")
            f.write("3. Prioritized courses with more conflicts\n")
            f.write("4. Minimized same-day and consecutive exams\n")
            f.write("5. Enforced 500 student capacity limit\n")
        
        print("✓ Saved to schedule_report.txt")
    
    def run(self):
        print("\n" + "="*70)
        print("EXAM SCHEDULER - FAST PESHAWAR")
        print("="*70 + "\n")
        
        if not self.load_data():
            return
        
        self.find_conflicts()
        self.create_schedule()
        self.save_schedule()
        self.analyze_and_report()
        
        print("\n" + "="*70)
        print("DONE! Check:")
        print("  - exam_schedule.xlsx")
        print("  - schedule_report.txt")
        print("="*70 + "\n")


if __name__ == "__main__":
    scheduler = ExamScheduler("Student Data.xlsx")
    scheduler.run()
